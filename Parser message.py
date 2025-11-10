"""
Telegram Channel Parser Bot (Без API ключей)
Парсит публичные сообщения из телеграм канала в Excel таблицу
Использует web-scraping подход без необходимости API
"""

import asyncio
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import aiohttp
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

# ========== НАСТРОЙКИ ==========
# Канал для парсинга (публичный канал)
CHANNEL_USERNAME = '@findwork'  # Замените на имя канала (только публичные!)

# Имя файла Excel
EXCEL_FILE = 'telegram_messages.xlsx'

# Период парсинга (по умолчанию 1 год для получения большего количества сообщений)
PARSE_PERIOD_DAYS = 365

# Максимальное количество сообщений для парсинга (для ограничения)
MAX_MESSAGES = 100000

# Фильтр по ключевым словам (оставьте пустым для парсинга всех сообщений)
# Можно указать одно слово или несколько через запятую
FILTER_KEYWORDS = ''  # Например: 'python,django,flask' или 'вакансия'
# Если указано, будут сохраняться только сообщения, содержащие хотя бы одно из этих слов
# ================================


class TelegramChannelParser:
    def __init__(self, channel_username, excel_file, filter_keywords=''):
        self.channel_username = channel_username.replace('@', '')
        self.excel_file = excel_file
        self.base_url = f"https://t.me/s/{self.channel_username}"
        self.workbook = None
        self.worksheet = None
        self.session = None
        # Обработка фильтра ключевых слов
        if filter_keywords:
            self.keywords = [kw.strip().lower() for kw in filter_keywords.split(',') if kw.strip()]
        else:
            self.keywords = []

    def matches_filter(self, text):
        """Проверка, соответствует ли текст фильтру ключевых слов"""
        if not self.keywords:
            return True  # Если фильтр не задан, пропускаем все сообщения

        text_lower = text.lower()
        for keyword in self.keywords:
            if keyword in text_lower:
                return True
        return False

    def init_excel(self):
        """Инициализация Excel файла"""
        if os.path.exists(self.excel_file):
            print(f"Загрузка существующего файла {self.excel_file}")
            self.workbook = load_workbook(self.excel_file)
            self.worksheet = self.workbook.active
        else:
            print(f"Создание нового файла {self.excel_file}")
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Messages"

            # Заголовки таблицы
            headers = ['ID', 'Дата', 'Время', 'Автор', 'Текст сообщения', 'Просмотры', 'Ссылка']
            self.worksheet.append(headers)

            # Форматирование заголовков
            for cell in self.worksheet[1]:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Установка ширины столбцов
            self.worksheet.column_dimensions['A'].width = 10
            self.worksheet.column_dimensions['B'].width = 12
            self.worksheet.column_dimensions['C'].width = 10
            self.worksheet.column_dimensions['D'].width = 20
            self.worksheet.column_dimensions['E'].width = 60
            self.worksheet.column_dimensions['F'].width = 12
            self.worksheet.column_dimensions['G'].width = 40

            self.save_excel()

    def save_excel(self):
        """Сохранение Excel файла"""
        self.workbook.save(self.excel_file)
        print(f"Файл {self.excel_file} сохранен")

    def message_exists(self, message_id):
        """Проверка существования сообщения в таблице"""
        for row in self.worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == message_id:
                return True
        return False

    def parse_message_date(self, date_element):
        """Парсинг даты сообщения"""
        try:
            if date_element and 'datetime' in date_element.attrs:
                date_str = date_element['datetime']
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                # Преобразуем в naive datetime (без timezone)
                return dt.replace(tzinfo=None)
        except:
            pass
        return None

    def parse_views(self, views_element):
        """Парсинг количества просмотров"""
        try:
            if views_element:
                views_text = views_element.get_text(strip=True)
                # Удаляем буквы K, M и конвертируем
                views_text = views_text.replace('K', '000').replace('M', '000000').replace(',', '')
                return int(''.join(filter(str.isdigit, views_text)))
        except:
            pass
        return 0

    async def fetch_page(self, url):
        """Получение HTML страницы"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        async with self.session.get(url, headers=headers) as response:
            if response.status == 200:
                return await response.text()
            else:
                print(f"Ошибка при получении страницы: {response.status}")
                return None

    async def parse_messages_from_page(self, html, start_date):
        """Парсинг сообщений со страницы"""
        soup = BeautifulSoup(html, 'html.parser')
        messages = soup.find_all('div', class_='tgme_widget_message')

        parsed_messages = []

        print(f"Найдено HTML блоков сообщений: {len(messages)}")

        for msg in messages:
            try:
                # Получаем ID сообщения
                message_link = msg.get('data-post', '')
                if not message_link:
                    # Пробуем альтернативный способ
                    link_elem = msg.find('a', class_='tgme_widget_message_date')
                    if link_elem and 'href' in link_elem.attrs:
                        message_link = link_elem['href']

                if '/' in message_link:
                    message_id = int(message_link.split('/')[-1])
                else:
                    continue

                # Проверяем, не добавлено ли уже
                if self.message_exists(message_id):
                    continue

                # Получаем дату
                date_element = msg.find('time', class_='datetime')
                if not date_element:
                    date_element = msg.find('time')

                msg_date = self.parse_message_date(date_element)

                if not msg_date:
                    # Пробуем получить из ссылки
                    continue

                # НЕ проверяем период здесь - будем проверять при добавлении
                # Это позволит парсить все страницы, а не останавливаться на первом старом сообщении

                date_str = msg_date.strftime('%d.%m.%Y')
                time_str = msg_date.strftime('%H:%M:%S')

                # Получаем автор (для каналов обычно название канала)
                author_element = msg.find('div', class_='tgme_widget_message_author')
                if not author_element:
                    author_element = msg.find('a', class_='tgme_widget_message_owner_name')
                author = author_element.get_text(strip=True) if author_element else self.channel_username

                # Получаем текст сообщения
                text_element = msg.find('div', class_='tgme_widget_message_text')
                text = ''
                if text_element:
                    # Получаем весь текст, включая вложенные элементы
                    text = text_element.get_text('\n', strip=True)

                # Проверяем наличие медиа
                if not text:
                    photo = msg.find('a', class_='tgme_widget_message_photo_wrap')
                    video = msg.find('video', class_='tgme_widget_message_video')
                    doc = msg.find('div', class_='tgme_widget_message_document')
                    if photo:
                        text = '[Фото]'
                    elif video:
                        text = '[Видео]'
                    elif doc:
                        text = '[Документ]'
                    else:
                        text = '[Медиа]'

                # Применяем фильтр по ключевым словам
                if not self.matches_filter(text):
                    continue  # Пропускаем сообщение, если оно не соответствует фильтру

                # Проверяем период - теперь ПОСЛЕ парсинга всех данных
                if msg_date < start_date:
                    continue  # Пропускаем старые сообщения

                # Получаем просмотры
                views_element = msg.find('span', class_='tgme_widget_message_views')
                views = self.parse_views(views_element)

                # Формируем ссылку
                link = f"https://t.me/{self.channel_username}/{message_id}"

                parsed_messages.append({
                    'id': message_id,
                    'date': date_str,
                    'time': time_str,
                    'author': author,
                    'text': text[:500] if len(text) > 500 else text,  # Ограничиваем длину
                    'views': views,
                    'link': link,
                    'datetime': msg_date
                })

            except Exception as e:
                print(f"Ошибка парсинга сообщения: {e}")
                continue

        return parsed_messages

    async def parse_historical_messages(self, days=365):
        """Парсинг исторических сообщений"""
        print(f"\n{'='*60}")
        print(f"Начало парсинга сообщений за последние {days} дней")
        if self.keywords:
            print(f"🔍 Активен фильтр по словам: {', '.join(self.keywords)}")
        print(f"{'='*60}\n")

        start_date = datetime.now() - timedelta(days=days)
        print(f"Парсинг канала: @{self.channel_username}")
        print(f"Парсинг сообщений с {start_date.strftime('%d.%m.%Y %H:%M:%S')}")
        print(f"URL: {self.base_url}\n")

        message_count = 0
        all_messages = []
        filtered_count = 0  # Счетчик отфильтрованных сообщений

        try:
            # Парсим основную страницу
            html = await self.fetch_page(self.base_url)
            if not html:
                print("Ошибка: Не удалось получить страницу канала")
                print("Убедитесь, что канал публичный и имя указано правильно")
                return

            messages = await self.parse_messages_from_page(html, start_date)
            all_messages.extend(messages)

            print(f"Найдено сообщений на первой странице: {len(messages)}")

            # Пытаемся получить больше сообщений через before параметр
            if messages:
                oldest_id = min(msg['id'] for msg in messages)
                empty_pages = 0  # Счетчик пустых страниц

                # Парсим предыдущие страницы (увеличиваем до 50 итераций)
                for i in range(50):
                    url = f"{self.base_url}?before={oldest_id}"
                    print(f"\nЗагрузка страницы {i+2}, URL: {url}")
                    html = await self.fetch_page(url)

                    if not html:
                        print(f"Не удалось загрузить страницу {i+2}")
                        break

                    new_messages = await self.parse_messages_from_page(html, start_date)

                    if not new_messages:
                        empty_pages += 1
                        print(f"⚠ Пустая страница {i+2} (пустых подряд: {empty_pages})")

                        # Останавливаемся только после 3 пустых страниц подряд
                        if empty_pages >= 3:
                            print(f"Остановка после {empty_pages} пустых страниц подряд")
                            break

                        # Пробуем получить следующую страницу
                        # Используем ID из HTML, если есть
                        soup = BeautifulSoup(html, 'html.parser')
                        msgs = soup.find_all('div', class_='tgme_widget_message')
                        if msgs:
                            last_msg = msgs[-1]
                            link = last_msg.get('data-post', '')
                            if '/' in link:
                                oldest_id = int(link.split('/')[-1])
                                continue
                        break
                    else:
                        empty_pages = 0  # Сбрасываем счетчик

                    all_messages.extend(new_messages)
                    oldest_id = min(msg['id'] for msg in new_messages)

                    print(f"✓ Загружено сообщений: {len(all_messages)} (страница {i+2}, новых: {len(new_messages)})")

                    # Проверка на лимит
                    if len(all_messages) >= MAX_MESSAGES:
                        print(f"⚠ Достигнут лимит в {MAX_MESSAGES} сообщений")
                        break

                    # Небольшая задержка между запросами
                    await asyncio.sleep(2)

            # Сортируем по дате (от старых к новым)
            all_messages.sort(key=lambda x: x['datetime'])

            # Добавляем в Excel
            for msg in all_messages:
                row = [msg['id'], msg['date'], msg['time'], msg['author'],
                       msg['text'], msg['views'], msg['link']]
                self.worksheet.append(row)
                message_count += 1

                print(f"Добавлено сообщение ID: {msg['id']} от {msg['date']} {msg['time']}")

                # Сохраняем каждые 50 сообщений
                if message_count % 50 == 0:
                    self.save_excel()

            # Финальное сохранение
            self.save_excel()
            print(f"\n{'='*60}")
            print(f"Парсинг завершен! Всего обработано: {message_count} сообщений")
            if self.keywords:
                print(f"🔍 Фильтр по словам [{', '.join(self.keywords)}] применен")
                print(f"📊 Сообщений соответствует фильтру: {message_count}")
            print(f"{'='*60}\n")

        except Exception as e:
            print(f"Ошибка при парсинге: {e}")
            import traceback
            traceback.print_exc()

    async def run(self):
        """Основной метод запуска бота"""
        try:
            print("Инициализация парсера...")

            # Создаем aiohttp сессию
            self.session = aiohttp.ClientSession()

            # Инициализация Excel
            self.init_excel()

            # Парсинг сообщений
            await self.parse_historical_messages(days=PARSE_PERIOD_DAYS)

        except KeyboardInterrupt:
            print("\n\nОстановка парсера...")
        except Exception as e:
            print(f"Критическая ошибка: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if self.workbook:
                self.save_excel()
            if self.session:
                await self.session.close()
            print("Парсер остановлен.")



async def main():
    """Точка входа в программу"""
    print("""
    ╔══════════════════════════════════════════════════════════╗
    ║     Telegram Channel Parser Bot (БЕЗ API)                ║
    ║     Парсер публичных каналов через Web                   ║
    ╚══════════════════════════════════════════════════════════╝
    """)

    # Проверка настроек
    if CHANNEL_USERNAME == '@your_channel':
        print("ОШИБКА: Необходимо указать имя канала для парсинга!")
        print("Откройте файл и замените '@your_channel' на имя канала")
        print("Например: '@durov' или '@python_job'")
        return

    print(f"Канал для парсинга: {CHANNEL_USERNAME}")
    print(f"Период: {PARSE_PERIOD_DAYS} дней")
    print(f"Выходной файл: {EXCEL_FILE}")
    print(f"Максимум сообщений: {MAX_MESSAGES}")
    if FILTER_KEYWORDS:
        print(f"Фильтр по словам: {FILTER_KEYWORDS}")
    print()

    # Создание и запуск парсера
    parser = TelegramChannelParser(
        channel_username=CHANNEL_USERNAME,
        excel_file=EXCEL_FILE,
        filter_keywords=FILTER_KEYWORDS
    )

    await parser.run()


if __name__ == '__main__':
    asyncio.run(main())



